#!/usr/bin/env python3
"""Generate Confluence markdown report and Excel from benchmark results.

Usage:
  python3 generate_report.py <results_dir>                     # single run
  python3 generate_report.py <dir1> <dir2> ... -o <output_dir> # merge multiple runs

All device, model, and kernel information is auto-detected from:
  - benchmark_config.json  (stage configs, model path, device info)
  - */output_stdout.txt     (per-stage benchmark output)
  - logcat_mnnjni.txt       (OpenCL kernel profiling, optional)

No hardcoded device names, model names, or kernel lists.
"""
import json, os, re, sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--break-system-packages", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

# ── Excel styling ──
_HDR_FONT  = Font(bold=True, size=11, color="FFFFFF")
_HDR_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_SUB_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SUB_FONT  = Font(bold=True, size=11)
_GREEN     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED       = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_BOLD  = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
_YELLOW    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_BORDER    = Border(*(Side('thin'),) * 4)
_TITLE_FNT = Font(bold=True, size=14)
_SEC_FNT   = Font(bold=True, size=12)
_NOTE_FNT  = Font(italic=True, size=10, color="666666")
_WARN_FNT  = Font(bold=True, color="FF0000")
_NUM_FMT   = '#,##0.0'
_INT_FMT   = '#,##0'
_PCT_FMT   = '0.0%'


def _hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.border = _HDR_FONT, _HDR_FILL, _BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def _bdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(row=row, column=c).border = _BORDER

def _bold(ws, row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).border = _BORDER

def _sub_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _SUB_FILL
        cell.font = _SUB_FONT
        cell.border = _BORDER

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_row(ws, row, values, fmt=None):
    """Write a row of values with optional number formatting."""
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = _BORDER
        if fmt and isinstance(v, (int, float)):
            cell.number_format = fmt

def _num_cell(ws, row, col, value, fmt=_NUM_FMT):
    """Write a single number cell with formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _BORDER
    if isinstance(value, (int, float)):
        cell.number_format = fmt
    return cell


# =====================================================================
# 0.  Parse arguments
# =====================================================================
input_dirs = []
output_dir = None

args = sys.argv[1:]
i = 0
while i < len(args):
    if args[i] == "-o" and i + 1 < len(args):
        output_dir = Path(args[i + 1])
        i += 2
    else:
        input_dirs.append(Path(args[i]))
        i += 1

if not input_dirs:
    print("Usage: python3 generate_report.py <results_dir> [<dir2> ...] [-o <output_dir>]")
    sys.exit(1)

if output_dir is None:
    output_dir = input_dirs[0]

output_dir.mkdir(parents=True, exist_ok=True)
print(f"Input dirs: {[str(d) for d in input_dirs]}")
print(f"Output dir: {output_dir}")


# =====================================================================
# 1.  Load configs and discover stages from all input dirs
# =====================================================================
all_bench_cfgs = []
for d in input_dirs:
    cfg_path = d / "benchmark_config.json"
    if cfg_path.exists():
        all_bench_cfgs.append(json.loads(cfg_path.read_text()))

bench_cfg = all_bench_cfgs[0] if all_bench_cfgs else {}

model_cfg  = bench_cfg.get("model", {})
device_cfg = bench_cfg.get("device", {})
inputs_cfg = bench_cfg.get("inputs", {})

model_name = model_cfg.get("model_type", "unknown")
model_dir  = model_cfg.get("converted_mnn_dir", "")
if model_dir:
    dirname = Path(model_dir).name
    if dirname:
        model_name = dirname

device_id  = device_cfg.get("adb_device_id", "unknown")
vlm_prompt = inputs_cfg.get("vlm_prompt_template", "")
images     = inputs_cfg.get("images", [])
image_desc = Path(images[0]).name if images else ""

# Load device info (collected from ADB at runtime)
device_info = {}
for d in input_dirs:
    di_path = d / "device_info.json"
    if di_path.exists():
        device_info = json.loads(di_path.read_text())
        print(f"Loaded device info: {device_info.get('brand', '?')} {device_info.get('model', '?')}")
        break

# Merge stage configs from all configs
stage_configs = {}
for cfg in all_bench_cfgs:
    for s in cfg.get("stages", []):
        if s.get("enabled", True):
            stage_configs[s["name"]] = s


# =====================================================================
# 2.  Parse each stage's stdout from all input dirs
# =====================================================================
def _extract(pattern, text, group=1, as_float=True):
    m = re.search(pattern, text)
    if not m:
        return None
    v = m.group(group)
    return float(v) if as_float else v

def parse_stage_stdout(stdout):
    s = {}
    s["backend"]      = _extract(r"Backend\s+│\s+(\w+)", stdout, as_float=False) or "?"
    s["mllm_backend"] = _extract(r"MLLM backend\s+│\s+(\w+)", stdout, as_float=False) or "?"
    s["mllm_detail"]  = _extract(r"MLLM backend\s+│\s+(.+?)║", stdout, as_float=False) or ""
    s["threads"]      = _extract(r"(?:Threads|thread_num)\s+│\s+(\d+)", stdout)
    s["precision"]    = _extract(r"Precision\s+│\s+(\w+)", stdout, as_float=False) or "?"

    s["vision_min"]   = _extract(r"Vision encode\s+ms\s+│\s+([\d.]+)", stdout)
    s["vision_avg"]   = _extract(r"Vision encode\s+ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["vision_max"]   = _extract(r"Vision encode\s+ms\s+│\s+[\d.]+\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)

    s["prefill_min"]  = _extract(r"LLM forward\s+ms\s+│\s+([\d.]+)", stdout) or _extract(r"LLM prefill\s+ms\s+│\s+([\d.]+)", stdout)
    s["prefill_avg"]  = _extract(r"LLM forward\s+ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout) or _extract(r"LLM prefill\s+ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)

    s["ttft_avg"]             = _extract(r"TTFT.*ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["embed_overhead_avg"]   = _extract(r"Embed\+overhead\s+ms\s+│\s+│\s+([\d.]+)", stdout)
    s["decode_ms_avg"]        = _extract(r"Decode\s+\(total\)\s+ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["decode_tok_min"]       = _extract(r"Decode  throughput tok/s\s+│\s+([\d.]+)", stdout)
    s["decode_tok_avg"]       = _extract(r"Decode  throughput tok/s\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["decode_tok_max"]       = _extract(r"Decode  throughput tok/s\s+│\s+[\d.]+\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["prefill_tok_avg"]      = _extract(r"Prefill throughput tok/s\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["tpot"]                 = _extract(r"Avg TPOT.*:\s+([\d.]+)", stdout)
    s["e2e_avg"]              = _extract(r"End-to-end\s+ms\s+│\s+[\d.]+\s+│\s+([\d.]+)", stdout)
    s["load_ms"]              = _extract(r"Model load time\s*:\s+([\d.]+)", stdout)

    gen_tokens = re.findall(r"│\s*\d+\s*│.*?│\s*\d+/(\d+)\s*│", stdout)
    if gen_tokens:
        s["avg_gen_tokens"] = sum(int(x) for x in gen_tokens) / len(gen_tokens)

    s["model_output"] = ""
    m = re.search(r"---BEGIN---\n(.*?)\n---END---", stdout, re.DOTALL)
    if m:
        s["model_output"] = m.group(1).strip()
    s["output_status"] = _extract(r"status=(\w+)", stdout, as_float=False) or "?"
    s["output_tokens"]  = _extract(r"gen_seq_len=(\d+)", stdout)

    return s


stages = {}
for d in input_dirs:
    # Find all output_stdout.txt files recursively
    stdout_files = sorted(d.rglob("output_stdout.txt"))
    for sf in stdout_files:
        # Determine stage name from directory structure.
        # Possible layouts:
        #   _combined/<ts>/stage_name/model/output_stdout.txt
        #   stage_name/<ts>/model/output_stdout.txt
        #   stage_name/model/output_stdout.txt
        # Strategy: walk from the input dir, skip timestamps and the model dir
        # (the dir containing output_stdout.txt), take the first meaningful name.
        rel_parts = sf.relative_to(d).parts[:-1]  # drop "output_stdout.txt"

        # The model dir is always the immediate parent of output_stdout.txt.
        # Walk ancestor dirs (excluding model dir) looking for a non-timestamp,
        # non-special directory name to use as stage name.
        # Also check the input dir's own ancestry for stage names.
        candidate_parts = rel_parts[:-1] if len(rel_parts) > 1 else []

        stage_name = None
        for p in candidate_parts:
            if p in ("device_tmp", "_combined"):
                continue
            if re.match(r'^\d{8}_\d{6}$', p):
                continue
            stage_name = p
            break

        if not stage_name:
            # Walk up from input dir: cpu_best/20260414_... → cpu_best
            for ancestor in [d] + list(d.parents):
                name = ancestor.name
                if not name or name in ("benchmark_results", ".", "_combined", "device_tmp"):
                    continue
                if re.match(r'^\d{8}_\d{6}$', name):
                    continue
                stage_name = name
                break

        if stage_name and stage_name not in stages:
            stdout = sf.read_text()
            stages[stage_name] = parse_stage_stdout(stdout)

print(f"Parsed {len(stages)} stages: {list(stages.keys())}")


# =====================================================================
# 3.  Parse kernel profiling (auto-detect phases: prefill/decode)
#     with artifact detection for While0-style profiling overhead
# =====================================================================
ARTIFACT_RATIO = 10  # if kernel avg > ARTIFACT_RATIO * module total, flag it

def parse_kernel_profiling(logcat_path):
    """Parse kernel profiling from logcat with profiling artifact detection."""
    total_re  = re.compile(r'total kernel time = (\d+)\s+us')
    kernel_re = re.compile(r'kernel time = (\d+)\s+us (.+)')

    modules = []
    cur_kernels = defaultdict(int)
    cur_kernel_counts = defaultdict(int)
    cur_kernel_list = []

    with open(logcat_path) as f:
        for line in f:
            if "MNNJNI" not in line:
                continue
            m = total_re.search(line)
            if m:
                modules.append({
                    "total_us": int(m.group(1)),
                    "kernels": dict(cur_kernels),
                    "kernel_counts": dict(cur_kernel_counts),
                    "kernel_list": list(cur_kernel_list),
                })
                cur_kernels = defaultdict(int)
                cur_kernel_counts = defaultdict(int)
                cur_kernel_list = []
                continue
            if "total kernel" in line:
                continue
            m = kernel_re.search(line)
            if m:
                name, t = m.group(2).strip(), int(m.group(1))
                cur_kernels[name] += t
                cur_kernel_counts[name] += 1
                cur_kernel_list.append((name, t))

    if len(modules) < 5:
        return None

    # ── Detect profiling artifacts ──
    # Compute a reference time: median of all "large" module totals.
    # A single kernel call exceeding ARTIFACT_RATIO * this median is flagged
    # as profiling overhead (e.g. While0 in prefill accumulating GPU sync time).
    # We cannot compare against the module's own total because the module total
    # already includes the artifact value (it's a sum of all kernels).
    large_mod_totals = sorted([m["total_us"] for m in modules if m["total_us"] > 5000])
    if large_mod_totals:
        median_mod_total = large_mod_totals[len(large_mod_totals) // 2]
    else:
        median_mod_total = 50000  # fallback

    artifacts = {}  # kernel_name -> {"flagged_us": total, "occurrences": count}
    clean_modules = []

    for mod in modules:
        mod_total = mod["total_us"]
        mod_clean_list = []
        has_artifact = False

        # Check individual kernel entries for values far exceeding
        # what any real GPU kernel should take
        for kname, ktime in mod["kernel_list"]:
            if ktime > ARTIFACT_RATIO * median_mod_total:
                if kname not in artifacts:
                    artifacts[kname] = {"flagged_us": 0, "occurrences": 0,
                                        "example_kernel_us": ktime,
                                        "example_module_us": mod_total}
                artifacts[kname]["flagged_us"] += ktime
                artifacts[kname]["occurrences"] += 1
                has_artifact = True
            else:
                mod_clean_list.append((kname, ktime))

        # Rebuild clean kernel aggregation from clean list
        clean_k = defaultdict(int)
        clean_c = defaultdict(int)
        for kname, ktime in mod_clean_list:
            clean_k[kname] += ktime
            clean_c[kname] += 1

        # Recalculate module total from clean kernels only
        clean_total = sum(clean_k.values()) if has_artifact else mod_total

        clean_modules.append({
            "total_us": clean_total,
            "original_total_us": mod_total,
            "kernels": dict(clean_k),
            "kernel_counts": dict(clean_c),
            "kernel_list": mod_clean_list,
            "has_artifact": has_artifact,
        })

    if artifacts:
        print(f"  Profiling artifacts detected:")
        for kname, info in artifacts.items():
            print(f"    {kname}: {info['occurrences']} occurrences, "
                  f"total flagged {info['flagged_us']/1e6:.1f}M us "
                  f"(example: kernel={info['example_kernel_us']} us vs module={info['example_module_us']} us)")

    # ── Auto-detect phase boundaries ──
    large_times = sorted([m["total_us"] for m in clean_modules if m["total_us"] > 5000], reverse=True)
    if large_times:
        median_large = large_times[len(large_times) // 2]
        prefill_threshold = median_large * 3
    else:
        prefill_threshold = 100000

    # Group modules into steps (large + trailing small)
    phase_steps = []
    i = 0
    while i < len(clean_modules):
        if clean_modules[i]["total_us"] > 5000:
            merged_kernels = defaultdict(int)
            merged_counts = defaultdict(int)
            merged_list = []
            merged_total = clean_modules[i]["total_us"]
            for kn, kt in clean_modules[i]["kernels"].items():
                merged_kernels[kn] += kt
            for kn, kc in clean_modules[i].get("kernel_counts", {}).items():
                merged_counts[kn] += kc
            merged_list.extend(clean_modules[i]["kernel_list"])
            has_art = clean_modules[i].get("has_artifact", False)
            j = i + 1
            while j < len(clean_modules) and clean_modules[j]["total_us"] <= 5000:
                merged_total += clean_modules[j]["total_us"]
                for kn, kt in clean_modules[j]["kernels"].items():
                    merged_kernels[kn] += kt
                for kn, kc in clean_modules[j].get("kernel_counts", {}).items():
                    merged_counts[kn] += kc
                merged_list.extend(clean_modules[j]["kernel_list"])
                has_art = has_art or clean_modules[j].get("has_artifact", False)
                j += 1

            label = "prefill" if merged_total > prefill_threshold else "decode"
            phase_steps.append((label, merged_total, dict(merged_kernels),
                                merged_list, dict(merged_counts), has_art))
            i = j
        else:
            i += 1

    # Build phase aggregation
    phase_data = {}
    for phase_name in ["all", "prefill", "decode"]:
        agg = defaultdict(lambda: {"us": 0, "count": 0})
        for label, total, kdict, _, kcounts, _ in phase_steps:
            if phase_name != "all" and label != phase_name:
                continue
            for kn, kt in kdict.items():
                agg[kn]["us"] += kt
                agg[kn]["count"] += kcounts.get(kn, 1)
        phase_data[phase_name] = dict(agg)

    prefill_steps = [(t, k, kl) for l, t, k, kl, _, _ in phase_steps if l == "prefill"]
    decode_steps  = [(t, k, kl) for l, t, k, kl, _, _ in phase_steps if l == "decode"]

    return {
        "modules": clean_modules,
        "phase_steps": phase_steps,
        "phase_data": phase_data,
        "prefill_steps": prefill_steps,
        "decode_steps": decode_steps,
        "prefill_total_us": sum(t for t, _, _ in prefill_steps),
        "decode_total_us": sum(t for t, _, _ in decode_steps),
        "grand_total_us": sum(t for _, t, _, _, _, _ in phase_steps),
        "artifacts": artifacts,
    }


# Try all input dirs for kernel profiling
kernel_data = None
for d in input_dirs:
    for fname in ["logcat_mnnjni.txt", "logcat_capture.txt"]:
        logcat_path = d / fname
        if logcat_path.exists() and logcat_path.stat().st_size > 100000:
            print(f"Parsing kernel profiling from {logcat_path}...")
            kernel_data = parse_kernel_profiling(logcat_path)
            if kernel_data:
                print(f"  {len(kernel_data['modules'])} module executions, "
                      f"{len(kernel_data['prefill_steps'])} prefill, "
                      f"{len(kernel_data['decode_steps'])} decode steps")
                break
    if kernel_data:
        break


# =====================================================================
# Helper: compute wall vs GPU comparison data
# =====================================================================
def _get_wall_gpu_comparison(stages, kernel_data):
    """Return dict with wall-clock vs GPU kernel comparison metrics."""
    if not kernel_data or not kernel_data["decode_steps"]:
        return None

    decode_steps = kernel_data["decode_steps"]
    avg_decode_kernel_ms = kernel_data["decode_total_us"] / len(decode_steps) / 1000
    gpu_toks = 1000 / avg_decode_kernel_ms if avg_decode_kernel_ms > 0 else 0

    # Find the opencl stage (non-kernel-profiling one)
    ocl_stage = None
    for sn, s in stages.items():
        if s.get("backend") == "opencl" and "kernel" not in sn.lower():
            ocl_stage = s
            break

    cmp = {
        "avg_decode_kernel_ms": avg_decode_kernel_ms,
        "gpu_only_toks": gpu_toks,
        "decode_steps_count": len(decode_steps),
    }

    if ocl_stage:
        if ocl_stage.get("tpot"):
            wall = ocl_stage["tpot"]
            cmp["wall_tpot"] = wall
            cmp["tpot_overhead_ms"] = wall - avg_decode_kernel_ms
            cmp["tpot_overhead_pct"] = (wall - avg_decode_kernel_ms) / wall * 100 if wall > 0 else 0
            cmp["wall_toks"] = 1000 / wall if wall > 0 else 0

        if ocl_stage.get("ttft_avg") and kernel_data["prefill_steps"]:
            wall_ttft = ocl_stage["ttft_avg"]
            gpu_ttft = kernel_data["prefill_total_us"] / len(kernel_data["prefill_steps"]) / 1000
            cmp["wall_ttft"] = wall_ttft
            cmp["gpu_ttft"] = gpu_ttft
            cmp["ttft_overhead_ms"] = wall_ttft - gpu_ttft
            cmp["ttft_overhead_pct"] = (wall_ttft - gpu_ttft) / wall_ttft * 100 if wall_ttft > 0 else 0

    return cmp


# =====================================================================
# 4.  Generate Confluence Markdown
# =====================================================================
METRICS = [
    ("LLM Backend",          "backend",           None),
    ("MLLM (Vision)",        "mllm_detail",       None),
    ("Precision",            "precision",          None),
    ("Model Load",           "load_ms",            "ms"),
    ("Vision Encode",        "vision_avg",         "ms"),
    ("LLM Forward",          "prefill_avg",        "ms"),
    ("Embed+Overhead",       "embed_overhead_avg", "ms"),
    ("TTFT",                 "ttft_avg",           "ms"),
    ("Decode Total",         "decode_ms_avg",      "ms"),
    ("Decode Throughput",    "decode_tok_avg",     "tok/s"),
    ("TPOT",                 "tpot",              "ms"),
    ("Prefill Throughput",   "prefill_tok_avg",    "tok/s"),
    ("End-to-End",           "e2e_avg",            "ms"),
    ("Output Status",        "output_status",      None),
    ("Output Tokens",        "output_tokens",      None),
]

# Higher-is-better metrics for highlighting
_HIGHER_BETTER = {"decode_tok_avg", "prefill_tok_avg"}
_LOWER_BETTER  = {"load_ms", "vision_avg", "prefill_avg", "embed_overhead_avg",
                   "ttft_avg", "decode_ms_avg", "tpot", "e2e_avg"}

def _fmt(v):
    if v is None:
        return "-"
    if isinstance(v, float):
        return f"{v:,.1f}"
    return str(v).strip()


md = []
md.append(f"# {model_name} Benchmark Report")
md.append("")

# ── Device Information ──
if device_info:
    md.append("## Device Information")
    md.append("")
    md.append("| Property | Value |")
    md.append("|----------|-------|")
    brand = (device_info.get("brand", "") or "").title()
    model = device_info.get("model", "")
    product = device_info.get("product", "")
    md.append(f"| **Device** | {brand} {model} ({product}) |")
    chipset = device_info.get("chipset", "")
    platform = device_info.get("platform", "")
    if chipset or platform:
        md.append(f"| **Chipset** | {chipset} ({platform}) |")
    gpu_model = device_info.get("gpu_model", "")
    if gpu_model:
        md.append(f"| **GPU** | {gpu_model} |")
    gpu_clock = device_info.get("gpu_clock_range", "")
    if gpu_clock:
        md.append(f"| **GPU Clock** | {gpu_clock} |")
    gpu_temp = device_info.get("gpu_temp", "")
    if gpu_temp:
        md.append(f"| **GPU Temp (at start)** | {gpu_temp}°C |")
    cores = device_info.get("cpu_cores", "")
    if cores:
        md.append(f"| **CPU Cores** | {cores} |")
    notable = device_info.get("cpu_notable_features", "")
    if notable:
        md.append(f"| **CPU Features** | {notable} |")
    android = device_info.get("android_version", "")
    sdk = device_info.get("sdk_version", "")
    if android:
        md.append(f"| **Android** | {android} (API {sdk}) |")
    md.append("")

# ── Benchmark Info ──
md.append("## Benchmark Info")
md.append("")
md.append("| | |")
md.append("|---|---|")
md.append(f"| **Model** | {model_name} |")
if not device_info:
    md.append(f"| **Device** | {device_id} |")
if image_desc:
    md.append(f"| **Image** | {image_desc} |")
if vlm_prompt:
    md.append(f"| **Prompt** | `{vlm_prompt}` |")
md.append("")

# ── Performance Comparison ──
if stages:
    md.append("## Performance Comparison")
    md.append("")

    stage_names = list(stages.keys())
    unit_col = "Unit"
    md.append("| Metric | " + unit_col + " | " + " | ".join(stage_names) + " | Winner |")
    md.append("|--------|------|" + "|".join(["------"] * len(stage_names)) + "|--------|")

    for label, key, unit in METRICS:
        vals = [stages[sn].get(key) for sn in stage_names]
        formatted = [_fmt(v) for v in vals]

        # Determine winner
        winner = ""
        numeric_vals = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float))]
        if len(numeric_vals) >= 2:
            if key in _HIGHER_BETTER:
                best_idx = max(numeric_vals, key=lambda x: x[1])[0]
                winner = stage_names[best_idx]
                formatted[best_idx] = f"**{formatted[best_idx]}**"
            elif key in _LOWER_BETTER:
                best_idx = min(numeric_vals, key=lambda x: x[1])[0]
                winner = stage_names[best_idx]
                formatted[best_idx] = f"**{formatted[best_idx]}**"

        unit_str = unit or ""
        md.append(f"| {label} | {unit_str} | " + " | ".join(formatted) + f" | {winner} |")
    md.append("")

    # Best decode summary
    decode_avgs = {k: v.get("decode_tok_avg") or 0 for k, v in stages.items()}
    if any(v > 0 for v in decode_avgs.values()):
        best = max(decode_avgs, key=decode_avgs.get)
        md.append(f"> **Best decode throughput:** {best} at {decode_avgs[best]:.1f} tok/s")
    md.append("")

# ── Model Outputs (truncated to 500 chars) ──
md.append("## Model Outputs")
md.append("")
for sname, s in stages.items():
    output = s.get("model_output", "")
    if not output:
        continue
    md.append(f"### {sname}")
    tokens = int(s.get('output_tokens', 0) or 0)
    md.append(f"*Status: {s.get('output_status', '?')}, Tokens: {tokens}*")
    md.append("")
    if len(output) > 500:
        output = output[:500] + "\n... (truncated)"
    md.append("```")
    md.append(output)
    md.append("```")
    md.append("")

# ── Key Findings ──
md.append("## Key Findings")
md.append("")
findings = []
decode_vals = {sn: s.get("decode_tok_avg") for sn, s in stages.items() if s.get("decode_tok_avg")}
if len(decode_vals) >= 2:
    best = max(decode_vals, key=decode_vals.get)
    worst = min(decode_vals, key=decode_vals.get)
    if decode_vals[worst] > 0:
        ratio = decode_vals[best] / decode_vals[worst]
        findings.append(f"**Best decode:** {best} ({decode_vals[best]:.1f} tok/s) -- "
                        f"{ratio:.1f}x faster than {worst} ({decode_vals[worst]:.1f} tok/s)")
tpot_vals = {sn: s.get("tpot") for sn, s in stages.items() if s.get("tpot")}
if len(tpot_vals) >= 2:
    best_t = min(tpot_vals, key=tpot_vals.get)
    findings.append(f"**Lowest TPOT:** {best_t} ({tpot_vals[best_t]:.1f} ms/token)")
for i, f_ in enumerate(findings, 1):
    md.append(f"{i}. {f_}")
md.append("")

# ── GPU Kernel Analysis ──
if kernel_data:
    decode_steps  = kernel_data["decode_steps"]
    prefill_steps = kernel_data["prefill_steps"]
    decode_total_us  = kernel_data["decode_total_us"]
    prefill_total_us = kernel_data["prefill_total_us"]
    grand_total_us   = kernel_data["grand_total_us"]
    phase_data       = kernel_data["phase_data"]
    artifacts        = kernel_data.get("artifacts", {})

    md.append("## GPU Kernel Analysis")
    md.append("")

    # Wall vs GPU comparison
    cmp = _get_wall_gpu_comparison(stages, kernel_data)
    if cmp:
        md.append("### Wall-Clock vs GPU Kernel Time")
        md.append("")
        md.append("| Metric | Wall-Clock | GPU Kernel | Overhead |")
        md.append("|--------|-----------|-----------|----------|")
        if "wall_tpot" in cmp:
            md.append(f"| **TPOT (ms)** | {cmp['wall_tpot']:.1f} | "
                      f"{cmp['avg_decode_kernel_ms']:.1f} | "
                      f"{cmp['tpot_overhead_pct']:.0f}% ({cmp['tpot_overhead_ms']:.1f} ms) |")
            md.append(f"| **Decode (tok/s)** | {cmp['wall_toks']:.1f} | "
                      f"{cmp['gpu_only_toks']:.1f} | |")
        if "wall_ttft" in cmp:
            md.append(f"| **TTFT (ms)** | {cmp['wall_ttft']:.1f} | "
                      f"{cmp['gpu_ttft']:.1f} | "
                      f"{cmp['ttft_overhead_pct']:.0f}% ({cmp['ttft_overhead_ms']:.1f} ms) |")
        md.append("")
        md.append(f"*GPU kernel time is measured via Session_Debug profiling. "
                  f"The overhead represents CPU-side scheduling, memory transfers, and queue management.*")
        md.append("")

    # Phase summary
    md.append("### Phase Summary")
    md.append("")
    md.append("| Phase | GPU Time (ms) | Steps | Avg/Step (ms) |")
    md.append("|-------|-------------|-------|--------------|")
    if prefill_steps:
        avg_p = prefill_total_us / len(prefill_steps) / 1000
        md.append(f"| **Prefill** | {prefill_total_us/1000:,.1f} | {len(prefill_steps)} | {avg_p:,.1f} |")
    if decode_steps:
        avg_d = decode_total_us / len(decode_steps) / 1000
        md.append(f"| **Decode** | {decode_total_us/1000:,.1f} | {len(decode_steps)} | {avg_d:,.1f} |")
    md.append(f"| **Total** | {grand_total_us/1000:,.1f} | {len(kernel_data['phase_steps'])} | |")
    md.append("")

    # Decode kernels (primary interest)
    md.append("### Decode Kernels (Top 10)")
    md.append("")
    decode_data = phase_data.get("decode", {})
    if decode_data:
        decode_phase_total = sum(d["us"] for d in decode_data.values())
        sorted_dk = sorted(decode_data.items(), key=lambda x: x[1]["us"], reverse=True)[:10]
        md.append("| Kernel | Total (ms) | Count | Avg (us) | Share |")
        md.append("|--------|-----------|-------|---------|-------|")
        for kn, kd in sorted_dk:
            pct = kd["us"] / decode_phase_total * 100 if decode_phase_total > 0 else 0
            avg = kd["us"] / kd["count"] if kd["count"] else 0
            md.append(f"| {kn} | {kd['us']/1000:,.1f} | {kd['count']:,} | {avg:,.0f} | {pct:.1f}% |")
        md.append("")

    # Prefill kernels
    md.append("### Prefill Kernels (Top 10)")
    md.append("")
    prefill_data = phase_data.get("prefill", {})
    if prefill_data:
        prefill_phase_total = sum(d["us"] for d in prefill_data.values())
        sorted_pk = sorted(prefill_data.items(), key=lambda x: x[1]["us"], reverse=True)[:10]
        md.append("| Kernel | Total (ms) | Count | Avg (us) | Share |")
        md.append("|--------|-----------|-------|---------|-------|")
        for kn, kd in sorted_pk:
            pct = kd["us"] / prefill_phase_total * 100 if prefill_phase_total > 0 else 0
            avg = kd["us"] / kd["count"] if kd["count"] else 0
            md.append(f"| {kn} | {kd['us']/1000:,.1f} | {kd['count']:,} | {avg:,.0f} | {pct:.1f}% |")
        md.append("")

    # Artifact notes
    if artifacts:
        md.append("> **Note on profiling artifacts:** The following kernels had individual call times "
                  "exceeding 10x the module's total reported time, indicating Session_Debug GPU sync "
                  "overhead rather than real GPU execution time. These entries have been excluded from "
                  "the totals above:")
        md.append(">")
        for kname, info in artifacts.items():
            md.append(f"> - **{kname}**: {info['occurrences']} occurrences flagged, "
                      f"total {info['flagged_us']/1e6:.2f}M us "
                      f"(example: single call {info['example_kernel_us']:,} us "
                      f"vs module total {info['example_module_us']:,} us)")
        md.append("")

# ── Stage Configurations ──
if stage_configs:
    md.append("## Stage Configurations")
    md.append("")
    all_keys = set()
    for sc in stage_configs.values():
        all_keys.update(sc.keys())
    priority = ["name", "enabled", "backend", "precision", "memory", "power",
                "threads", "gpu_mode", "mllm_backend", "mllm_threads",
                "warmup_rounds", "measure_rounds", "max_gen_tokens",
                "enable_op_profile", "no_tuning"]
    ordered = [k for k in priority if k in all_keys] + sorted(all_keys - set(priority))

    cfg_names = list(stage_configs.keys())
    md.append("| Parameter | " + " | ".join(cfg_names) + " |")
    md.append("|-----------|" + "|".join(["---"] * len(cfg_names)) + "|")
    for key in ordered:
        vals = []
        for sn in cfg_names:
            v = stage_configs[sn].get(key, "-")
            if isinstance(v, (dict, list)):
                v = json.dumps(v)
            vals.append(str(v))
        md.append(f"| {key} | " + " | ".join(vals) + " |")
    md.append("")

# ── Methodology note ──
md.append("---")
md.append("")
md.append("*Profiling methodology: GPU kernel times are collected via MNN Session_Debug mode "
          "which inserts clFinish() barriers after each kernel for accurate per-kernel timing. "
          "This adds overhead that inflates wall-clock time during profiling runs (opencl_kernel_ops). "
          "Wall-clock metrics (TPOT, TTFT, tok/s) are taken from non-profiling runs (opencl_best). "
          "Kernels whose individual call times exceed 10x their module total are flagged as profiling "
          "artifacts and excluded from analysis.*")
md.append("")

md_path = output_dir / "confluence_report.md"
md_path.write_text("\n".join(md))
print(f"Saved: {md_path}")


# =====================================================================
# 5.  Generate Excel
# =====================================================================
wb = Workbook()

# ── Sheet 1: Summary ──
ws = wb.active
ws.title = "Summary"

# Title
ws.cell(1, 1, f"{model_name} Benchmark Report")
ws.cell(1, 1).font = _TITLE_FNT
stage_names = list(stages.keys())
ncols = max(3, len(stage_names) + 2)
ws.merge_cells(f"A1:{get_column_letter(ncols)}1")

row = 3
# Device info block in Excel
if device_info:
    ws.cell(row, 1, "Device Information")
    ws.cell(row, 1).font = _SEC_FNT
    row += 1
    dev_rows = [
        ("Device", f"{(device_info.get('brand','') or '').title()} {device_info.get('model','')}"),
        ("Chipset", f"{device_info.get('chipset','')} ({device_info.get('platform','')})"),
        ("GPU", device_info.get("gpu_model", "")),
        ("GPU Clock", device_info.get("gpu_clock_range", "")),
        ("CPU Cores", device_info.get("cpu_cores", "")),
        ("CPU Features", device_info.get("cpu_notable_features", "")),
        ("Android", f"{device_info.get('android_version','')} (API {device_info.get('sdk_version','')})"),
    ]
    for label, val in dev_rows:
        if val and val.strip() and val.strip() != "()":
            ws.cell(row, 1, label).font = Font(bold=True, size=10)
            ws.cell(row, 2, val.strip())
            ws.merge_cells(f"B{row}:{get_column_letter(ncols)}{row}")
            _bdr(ws, row, 2)
            row += 1
    row += 1
else:
    ws.cell(row, 1, f"Device: {device_id}")
    ws.cell(row, 1).font = Font(size=11, color="555555")
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    row += 2

if vlm_prompt:
    ws.cell(row, 1, "Prompt")
    ws.cell(row, 1).font = Font(bold=True, size=10)
    ws.cell(row, 2, vlm_prompt)
    ws.merge_cells(f"B{row}:{get_column_letter(ncols)}{row}")
    row += 1

row += 1
headers = ["Metric", "Unit"] + stage_names
for c, h in enumerate(headers, 1):
    ws.cell(row, c, h)
_hdr(ws, row, len(headers))

for label, key, unit in METRICS:
    row += 1
    ws.cell(row, 1, label)
    ws.cell(row, 2, unit or "")
    numeric_vals = []
    for ci, sn in enumerate(stage_names, 3):
        v = stages[sn].get(key)
        cell = ws.cell(row, ci)
        if isinstance(v, (int, float)):
            cell.value = round(v, 1) if isinstance(v, float) else v
            cell.number_format = _NUM_FMT
            numeric_vals.append((ci, v))
        else:
            cell.value = _fmt(v)
    _bdr(ws, row, len(headers))

    # Highlight best value
    if len(numeric_vals) >= 2:
        if key in _HIGHER_BETTER:
            best_col = max(numeric_vals, key=lambda x: x[1])[0]
            ws.cell(row, best_col).fill = _GREEN
        elif key in _LOWER_BETTER:
            best_col = min(numeric_vals, key=lambda x: x[1])[0]
            ws.cell(row, best_col).fill = _GREEN

_widths(ws, [22, 8] + [18] * len(stage_names))

# ── Sheet 2: GPU Kernel Profile (Decode) ──
if kernel_data and kernel_data["decode_steps"]:
    wd = wb.create_sheet("GPU Kernel Profile")
    cmp = _get_wall_gpu_comparison(stages, kernel_data)

    # Title
    wd.cell(1, 1, "GPU Kernel Profile - Decode Phase")
    wd.cell(1, 1).font = _TITLE_FNT
    wd.merge_cells("A1:F1")

    r = 3
    # Wall vs GPU summary section
    if cmp:
        wd.cell(r, 1, "Wall-Clock vs GPU Kernel Summary")
        wd.cell(r, 1).font = _SEC_FNT
        r += 1
        sum_headers = ["Metric", "Wall-Clock", "GPU Kernel", "Overhead"]
        for c, h in enumerate(sum_headers, 1):
            wd.cell(r, c, h)
        _hdr(wd, r, len(sum_headers))

        r += 1
        if "wall_tpot" in cmp:
            _write_row(wd, r, [
                "TPOT (ms/token)",
                round(cmp["wall_tpot"], 1),
                round(cmp["avg_decode_kernel_ms"], 1),
                f"{cmp['tpot_overhead_pct']:.0f}%"
            ])
            r += 1
            _write_row(wd, r, [
                "Decode (tok/s)",
                round(cmp["wall_toks"], 1),
                round(cmp["gpu_only_toks"], 1),
                ""
            ])
            r += 1
        if "wall_ttft" in cmp:
            _write_row(wd, r, [
                "TTFT (ms)",
                round(cmp["wall_ttft"], 1),
                round(cmp["gpu_ttft"], 1),
                f"{cmp['ttft_overhead_pct']:.0f}%"
            ])
            r += 1

        _write_row(wd, r, ["Decode steps profiled", len(kernel_data["decode_steps"]), "", ""])
        r += 2

    # Kernel table
    decode_data = kernel_data["phase_data"].get("decode", {})
    decode_phase_total = sum(d["us"] for d in decode_data.values())

    wd.cell(r, 1, f"Decode Kernels (Total: {decode_phase_total/1000:,.1f} ms)")
    wd.cell(r, 1).font = _SEC_FNT
    r += 1

    k_headers = ["Kernel Name", "Count", "Total (ms)", "Avg (us)", "% Share"]
    for c, h in enumerate(k_headers, 1):
        wd.cell(r, c, h)
    _hdr(wd, r, len(k_headers))

    sorted_dk = sorted(decode_data.items(), key=lambda x: x[1]["us"], reverse=True)
    for kn, kd in sorted_dk:
        r += 1
        pct = kd["us"] / decode_phase_total * 100 if decode_phase_total > 0 else 0
        avg = kd["us"] / kd["count"] if kd["count"] else 0
        wd.cell(r, 1, kn)
        _num_cell(wd, r, 2, kd["count"], _INT_FMT)
        _num_cell(wd, r, 3, round(kd["us"] / 1000, 1), _NUM_FMT)
        _num_cell(wd, r, 4, round(avg, 1), _NUM_FMT)
        pct_cell = wd.cell(r, 5, round(pct, 1))
        pct_cell.number_format = '0.0"%"'
        _bdr(wd, r, len(k_headers))

        # Color code by share
        if pct > 20:
            wd.cell(r, 1).fill = _RED_BOLD
            wd.cell(r, 5).fill = _RED_BOLD
        elif pct > 5:
            wd.cell(r, 1).fill = _YELLOW
            wd.cell(r, 5).fill = _YELLOW

    # Total row
    r += 1
    wd.cell(r, 1, "TOTAL")
    _num_cell(wd, r, 3, round(decode_phase_total / 1000, 1), _NUM_FMT)
    wd.cell(r, 5, "100.0%")
    _bold(wd, r, len(k_headers))

    _widths(wd, [32, 10, 14, 14, 12])

# ── Sheet 3: Prefill Kernels ──
if kernel_data and kernel_data["prefill_steps"]:
    wp = wb.create_sheet("Prefill Kernels")
    artifacts = kernel_data.get("artifacts", {})

    wp.cell(1, 1, "GPU Kernel Profile - Prefill Phase")
    wp.cell(1, 1).font = _TITLE_FNT
    wp.merge_cells("A1:F1")

    r = 3
    # Artifact warning
    if artifacts:
        wp.cell(r, 1, "PROFILING ARTIFACT NOTE:")
        wp.cell(r, 1).font = _WARN_FNT
        r += 1
        for kname, info in artifacts.items():
            wp.cell(r, 1, f"{kname}: {info['occurrences']} call(s) had times exceeding 10x the "
                          f"module total (profiling sync overhead). These have been excluded from "
                          f"all totals below. Flagged total: {info['flagged_us']/1e6:.2f}M us.")
            wp.cell(r, 1).font = _NOTE_FNT
            wp.merge_cells(f"A{r}:F{r}")
            r += 1
        r += 1

    prefill_data = kernel_data["phase_data"].get("prefill", {})
    prefill_phase_total = sum(d["us"] for d in prefill_data.values())

    wp.cell(r, 1, f"Prefill Kernels (Total: {prefill_phase_total/1000:,.1f} ms, "
                  f"Steps: {len(kernel_data['prefill_steps'])})")
    wp.cell(r, 1).font = _SEC_FNT
    r += 1

    k_headers = ["Kernel Name", "Count", "Total (ms)", "Avg (us)", "% Share"]
    for c, h in enumerate(k_headers, 1):
        wp.cell(r, c, h)
    _hdr(wp, r, len(k_headers))

    sorted_pk = sorted(prefill_data.items(), key=lambda x: x[1]["us"], reverse=True)
    for kn, kd in sorted_pk:
        r += 1
        pct = kd["us"] / prefill_phase_total * 100 if prefill_phase_total > 0 else 0
        avg = kd["us"] / kd["count"] if kd["count"] else 0
        wp.cell(r, 1, kn)
        _num_cell(wp, r, 2, kd["count"], _INT_FMT)
        _num_cell(wp, r, 3, round(kd["us"] / 1000, 1), _NUM_FMT)
        _num_cell(wp, r, 4, round(avg, 1), _NUM_FMT)
        pct_cell = wp.cell(r, 5, round(pct, 1))
        pct_cell.number_format = '0.0"%"'
        _bdr(wp, r, len(k_headers))

        if pct > 20:
            wp.cell(r, 1).fill = _RED_BOLD
            wp.cell(r, 5).fill = _RED_BOLD
        elif pct > 5:
            wp.cell(r, 1).fill = _YELLOW
            wp.cell(r, 5).fill = _YELLOW

    r += 1
    wp.cell(r, 1, "TOTAL")
    _num_cell(wp, r, 3, round(prefill_phase_total / 1000, 1), _NUM_FMT)
    wp.cell(r, 5, "100.0%")
    _bold(wp, r, len(k_headers))

    _widths(wp, [32, 10, 14, 14, 12])

# ── Sheet 4: Decode Per-Step ──
if kernel_data and kernel_data["decode_steps"]:
    ws4 = wb.create_sheet("Decode Per-Step")
    decode_steps = kernel_data["decode_steps"]

    ws4.cell(1, 1, "Per-Step Decode Breakdown")
    ws4.cell(1, 1).font = _TITLE_FNT

    # Find top kernels across all decode steps
    kname_totals = defaultdict(int)
    for _, kdict, _ in decode_steps:
        for kn, kt in kdict.items():
            kname_totals[kn] += kt
    sorted_knames = sorted(kname_totals.keys(), key=lambda k: kname_totals[k], reverse=True)

    MAX_COLS = 8
    top_knames = sorted_knames[:MAX_COLS]
    has_other = len(sorted_knames) > MAX_COLS

    r = 3
    headers = ["Step", "Total (ms)"]
    for kn in top_knames:
        headers.append(f"{kn} (us)")
    headers.append("Top Kernel %")
    if has_other:
        headers.append("Other (us)")

    for c, h in enumerate(headers, 1):
        ws4.cell(r, c, h)
    _hdr(ws4, r, len(headers))

    for idx, (total_us, kdict, _) in enumerate(decode_steps[:300], 1):
        r += 1
        ws4.cell(r, 1, idx)
        _num_cell(ws4, r, 2, round(total_us / 1000, 2), '#,##0.00')
        accounted = 0
        for ki, kn in enumerate(top_knames, 3):
            v = kdict.get(kn, 0)
            _num_cell(ws4, r, ki, v, _INT_FMT)
            accounted += v
        pct_col = 3 + len(top_knames)
        top_val = kdict.get(top_knames[0], 0) if top_knames else 0
        pct_v = top_val / total_us * 100 if total_us > 0 else 0
        ws4.cell(r, pct_col, round(pct_v, 1))
        ws4.cell(r, pct_col).number_format = '0.0"%"'
        ws4.cell(r, pct_col).border = _BORDER
        if has_other:
            _num_cell(ws4, r, pct_col + 1, total_us - accounted, _INT_FMT)
        _bdr(ws4, r, len(headers))

    _widths(ws4, [8, 14] + [14] * len(top_knames) + [12] + ([14] if has_other else []))

# ── Sheet 5: Stage Configs ──
if stage_configs:
    wc = wb.create_sheet("Stage Configs")
    wc.cell(1, 1, "Stage Configuration Parameters")
    wc.cell(1, 1).font = _TITLE_FNT
    wc.merge_cells(f"A1:{get_column_letter(max(3, len(stage_configs)+1))}1")

    r = 3
    cfg_names = list(stage_configs.keys())
    ch = ["Parameter"] + cfg_names
    for c, h in enumerate(ch, 1):
        wc.cell(r, c, h)
    _hdr(wc, r, len(ch))

    all_keys = set()
    for sc in stage_configs.values():
        all_keys.update(sc.keys())
    priority = ["name", "enabled", "backend", "precision", "memory", "power",
                "threads", "gpu_mode", "mllm_backend", "mllm_precision",
                "mllm_threads", "warmup_rounds", "measure_rounds",
                "max_gen_tokens", "no_tuning", "enable_op_profile",
                "use_vlm_input", "session_mode", "hints"]
    ordered = [k for k in priority if k in all_keys] + sorted(all_keys - set(priority))

    for key in ordered:
        r += 1
        wc.cell(r, 1, key)
        for ci, sn in enumerate(cfg_names, 2):
            v = stage_configs[sn].get(key, "-")
            wc.cell(r, ci, json.dumps(v) if isinstance(v, (dict, list)) else str(v))
        _bdr(wc, r, len(ch))

    _widths(wc, [22] + [22] * len(cfg_names))


xlsx_path = output_dir / "full_report.xlsx"
wb.save(str(xlsx_path))
print(f"Saved: {xlsx_path}")
print("Done!")
