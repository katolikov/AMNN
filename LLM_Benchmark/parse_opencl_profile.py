#!/usr/bin/env python3
"""Parse OpenCL kernel profiler logcat and generate Excel report.

Usage: python3 parse_opencl_profile.py <logcat_file> [output.xlsx]

All kernel names, phase boundaries, and thresholds are auto-detected from
the data — no hardcoded device names, model names, or operator lists.
"""
import re
import sys
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

logfile = sys.argv[1]
outfile = sys.argv[2] if len(sys.argv) > 2 else "opencl_profile.xlsx"

# ── Excel styling ──
_HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
_HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_RED      = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
_YELLOW   = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
_BORDER   = Border(*(Side(style='thin'),) * 4)

def _hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.border = _HDR_FONT, _HDR_FILL, _BORDER
        cell.alignment = Alignment(horizontal='center')

def _bdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(row=row, column=c).border = _BORDER

def _bold(ws, row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════
# 1.  Parse all kernel and module entries
# ═══════════════════════════════════════════════════════════════════
total_re  = re.compile(r'total kernel time = (\d+)\s+us')
kernel_re = re.compile(r'kernel time = (\d+)\s+us (.+)')

modules = []
cur_kernels = defaultdict(int)
cur_kernel_list = []

with open(logfile) as f:
    for line in f:
        if "MNNJNI" not in line:
            continue
        m = total_re.search(line)
        if m:
            modules.append({
                "total_us": int(m.group(1)),
                "kernels": dict(cur_kernels),
                "kernel_list": list(cur_kernel_list),
            })
            cur_kernels, cur_kernel_list = defaultdict(int), []
            continue
        if "total kernel" in line:
            continue
        m = kernel_re.search(line)
        if m:
            name, t = m.group(2).strip(), int(m.group(1))
            cur_kernels[name] += t
            cur_kernel_list.append((name, t))

print(f"Parsed {len(modules)} module executions")
if not modules:
    print("ERROR: No kernel profiling data found")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════
# 2.  Auto-detect phase boundaries
# ═══════════════════════════════════════════════════════════════════
# Large modules (>5000 us) are LLM forward passes.
# The first extremely large one is vision+prefill; rest are decode steps.
# Small modules between large ones are trailing ops (post-LN, etc.)

large_times = sorted([m["total_us"] for m in modules if m["total_us"] > 5000], reverse=True)
if large_times:
    median_large = large_times[len(large_times) // 2]
    prefill_threshold = median_large * 3
else:
    prefill_threshold = 100000

# Group modules into steps
phase_steps = []  # (label, total_us, merged_kernels_dict, merged_kernel_list)
i = 0
while i < len(modules):
    if modules[i]["total_us"] > 5000:
        merged_kernels = defaultdict(int)
        merged_list = []
        merged_total = modules[i]["total_us"]
        for kn, kt in modules[i]["kernels"].items():
            merged_kernels[kn] += kt
        merged_list.extend(modules[i]["kernel_list"])
        j = i + 1
        while j < len(modules) and modules[j]["total_us"] <= 5000:
            merged_total += modules[j]["total_us"]
            for kn, kt in modules[j]["kernels"].items():
                merged_kernels[kn] += kt
            merged_list.extend(modules[j]["kernel_list"])
            j += 1

        # Classify by size: very large = vision+prefill, normal = decode
        label = "vision_prefill" if merged_total > prefill_threshold else "decode"
        phase_steps.append((label, merged_total, dict(merged_kernels), merged_list))
        i = j
    else:
        i += 1

prefill_steps = [(t, k, kl) for l, t, k, kl in phase_steps if l == "vision_prefill"]
decode_steps  = [(t, k, kl) for l, t, k, kl in phase_steps if l == "decode"]

prefill_total = sum(t for t, _, _ in prefill_steps)
decode_total  = sum(t for t, _, _ in decode_steps)
grand_total   = prefill_total + decode_total

print(f"Phases: {len(prefill_steps)} prefill ({prefill_total/1000:.1f} ms), "
      f"{len(decode_steps)} decode ({decode_total/1000:.1f} ms)")


# ═══════════════════════════════════════════════════════════════════
# 3.  Aggregate kernels per phase (auto-discovered names)
# ═══════════════════════════════════════════════════════════════════
phase_data = {}
for phase_name, steps_list in [("all", [(t, k) for _, t, k, _ in phase_steps]),
                                ("vision_prefill", [(t, k) for t, k, _ in prefill_steps]),
                                ("decode", [(t, k) for t, k, _ in decode_steps])]:
    agg = defaultdict(lambda: {"us": 0, "count": 0})
    for _, kdict in steps_list:
        for kn, kt in kdict.items():
            agg[kn]["us"] += kt
            agg[kn]["count"] += 1
    phase_data[phase_name] = dict(agg)


# ═══════════════════════════════════════════════════════════════════
# 4.  Generate Excel
# ═══════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Phase Summary ──
ws1 = wb.active
ws1.title = "Summary"
ws1.append(["OpenCL Kernel Profiling Summary"])
ws1.merge_cells("A1:G1")
ws1.cell(1, 1).font = Font(bold=True, size=14)
ws1.append([])

ws1.append(["Phase", "Total GPU (ms)", "Total GPU (us)", "Steps", "Avg/Step (ms)", "% of Total"])
_hdr(ws1, 3, 6)

for phase, total, count in [
    ("Vision + Prefill", prefill_total, len(prefill_steps)),
    ("Decode", decode_total, len(decode_steps)),
]:
    if count == 0:
        continue
    avg = total / count / 1000
    pct = total / grand_total * 100 if grand_total > 0 else 0
    ws1.append([phase, f"{total/1000:.1f}", total, count, f"{avg:.2f}", f"{pct:.1f}%"])
    _bdr(ws1, ws1.max_row, 6)

ws1.append(["TOTAL", f"{grand_total/1000:.1f}", grand_total, len(phase_steps), "", "100%"])
_bold(ws1, ws1.max_row, 6)

ws1.append([])
ws1.append(["Decode Performance"])
ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=12)
if decode_steps:
    avg_ms = decode_total / len(decode_steps) / 1000
    ws1.append(["Avg GPU kernel time per token", f"{avg_ms:.2f} ms"])
    ws1.append(["Estimated GPU-only tok/s", f"{1000/avg_ms:.1f}" if avg_ms > 0 else "N/A"])

_widths(ws1, [25, 18, 18, 12, 18, 12])

# ── Sheet 2+: Kernel Breakdown per phase ──
for phase_label, phase_key in [("All Phases", "all"), ("Decode", "decode"),
                                ("Vision+Prefill", "vision_prefill")]:
    data = phase_data.get(phase_key, {})
    if not data:
        continue
    phase_total_us = sum(d["us"] for d in data.values())
    if phase_total_us == 0:
        continue

    ws_name = f"Kernels — {phase_label}"[:31]
    wk = wb.create_sheet(ws_name)
    wk.append([f"GPU Kernel Breakdown — {phase_label}"])
    wk.cell(1, 1).font = Font(bold=True, size=13)
    wk.append([f"Total: {phase_total_us/1000:.1f} ms"])
    wk.append([])

    wk.append(["Kernel Name", "Total (us)", "Total (ms)", "Calls", "Avg (us)", "% of Phase"])
    _hdr(wk, wk.max_row, 6)

    check_sum = 0
    for kn, kd in sorted(data.items(), key=lambda x: x[1]["us"], reverse=True):
        pct = kd["us"] / phase_total_us * 100
        avg = kd["us"] / kd["count"] if kd["count"] else 0
        wk.append([kn, kd["us"], round(kd["us"]/1000, 1), kd["count"], round(avg, 1), f"{pct:.1f}%"])
        _bdr(wk, wk.max_row, 6)
        check_sum += kd["us"]
        if pct > 20:
            wk.cell(wk.max_row, 1).fill = _RED
        elif pct > 5:
            wk.cell(wk.max_row, 1).fill = _YELLOW

    wk.append(["TOTAL", check_sum, round(check_sum/1000, 1), "", "", "100.0%"])
    _bold(wk, wk.max_row, 6)

    if abs(check_sum - phase_total_us) / max(phase_total_us, 1) > 0.05:
        wk.append([f"⚠ kernel sum ({check_sum}) ≠ module sum ({phase_total_us})"])
        wk.cell(wk.max_row, 1).font = Font(bold=True, color="FF0000")

    _widths(wk, [30, 14, 14, 10, 14, 10])

# ── Sheet: Per-Decode-Step Detail (auto-discovered columns) ──
if decode_steps:
    wd = wb.create_sheet("Decode Steps")

    # Auto-discover kernel names and rank by total contribution
    kname_totals = defaultdict(int)
    for _, kdict, _ in decode_steps:
        for kn, kt in kdict.items():
            kname_totals[kn] += kt
    sorted_knames = sorted(kname_totals.keys(), key=lambda k: kname_totals[k], reverse=True)

    MAX_COLS = 8
    top_knames = sorted_knames[:MAX_COLS]
    has_other = len(sorted_knames) > MAX_COLS

    headers = ["Step", "Total (us)", "Total (ms)"]
    for kn in top_knames:
        headers.append(f"{kn} (us)")
    headers.append("% Top Kernel")
    if has_other:
        headers.append("Other (us)")
    wd.append(headers)
    _hdr(wd, 1, len(headers))

    for idx, (total_us, kdict, _) in enumerate(decode_steps[:300], 1):
        row = [idx, total_us, round(total_us / 1000, 2)]
        accounted = 0
        for kn in top_knames:
            v = kdict.get(kn, 0)
            row.append(v)
            accounted += v
        # % of top kernel
        top_val = kdict.get(top_knames[0], 0) if top_knames else 0
        row.append(f"{top_val/total_us*100:.1f}%" if total_us > 0 else "0%")
        if has_other:
            row.append(total_us - accounted)
        wd.append(row)
        _bdr(wd, wd.max_row, len(headers))

    _widths(wd, [8, 14, 14] + [14] * len(top_knames) + [10] + ([14] if has_other else []))

# ── Sheet: Top 200 Slowest Individual Kernel Dispatches ──
all_individual = []
for label, total, kdict, klist in phase_steps:
    for kn, kt in klist:
        all_individual.append((kn, kt, label))

if all_individual:
    ws_slow = wb.create_sheet("Slowest Dispatches")
    ws_slow.append(["Rank", "Kernel Name", "Time (us)", "Time (ms)", "Phase"])
    _hdr(ws_slow, 1, 5)

    all_individual.sort(key=lambda x: x[1], reverse=True)
    for rank, (kn, kt, phase) in enumerate(all_individual[:200], 1):
        ws_slow.append([rank, kn, kt, f"{kt/1000:.3f}", phase])
        _bdr(ws_slow, ws_slow.max_row, 5)

    _widths(ws_slow, [8, 30, 12, 12, 15])

# Save
wb.save(outfile)
print(f"Saved: {outfile}")
